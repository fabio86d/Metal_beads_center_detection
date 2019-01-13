import numpy as np
import openpyxl


# reads the Sheet 'Sheetname' (default = Sheet1) of an .xlsx file in the range specified by the tuples of 2 integers range_rows and range_cols
def read_xlsx_to_nparray(filename, range_rows, range_cols, sheetname = 'Sheet1', dtype = None):

    output = np.zeros((range_rows[1] - range_rows[0], range_cols[1] - range_cols[0]), dtype)

    wb = openpyxl.load_workbook(filename)

    ws = wb.get_sheet_by_name(sheetname)

    for i in range(range_rows[0], range_rows[1]):
        for j in range(range_cols[0], range_cols[1]):

            output[i - range_rows[0]][j - range_cols[0]] = ws.cell(row=i, column=j).value

    return output

# generates output_filename.xlsx and writes numpy array into its 'Sheet1' 
def write_nparray_to_xlsx(input_array, output_filename, sheetname = 'Sheet1'):

    wb = openpyxl.Workbook()

    ws = wb.create_sheet(index = 0, title = sheetname)

    if not np.ndim(input_array) == 1:

        for i in range(input_array.shape[0]):
            
            ws.append(tuple(input_array[i]))

    else: 
        
        for i in range(input_array.shape[0]):
            
            ws.cell(row = i+1, column = 1).value = input_array[i]

    wb.save(output_filename)



if __name__ == "__main__":

    filename = 'p_ct.xlsx'
    sheetname = 'Sheet2'
    range_rows = (2,5)
    range_cols = (2,5)
          
    output = read_xlsx_to_nparray(filename, range_rows, range_cols, sheetname)
    
    print(output)

    #write_nparray_to_xlsx(output, 'sample_src_point_modified.xlsx')

